"""
教育数据分析报告生成引擎
使用Jinja2模板系统生成PDF和Excel格式的教育分析报告
"""

import os
import json
import logging
from typing import Dict, List, Any, Optional
from datetime import datetime
from dataclasses import dataclass
from pathlib import Path

import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from jinja2 import Environment, FileSystemLoader, select_autoescape
import pdfkit
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

from ..models.edu_data_models import (
    EduReportRequest, EduReportMetadata, EduSubject,
    EduGradeLevel, EduDataPrivacyLevel
)
from ..config.edu_data_config import edu_config

logger = logging.getLogger(__name__)

# 设置matplotlib中文字体
plt.rcParams['font.sans-serif'] = ['SimHei', 'Arial Unicode MS', 'DejaVu Sans']
plt.rcParams['axes.unicode_minus'] = False


@dataclass
class ReportTemplate:
    """报告模板配置"""
    name: str
    template_file: str
    output_format: str
    charts_enabled: bool
    detailed_stats: bool


class EduReportGenerator:
    """教育报告生成器"""

    def __init__(self, template_dir: str = None, output_dir: str = None):
        self.template_dir = template_dir or edu_config.report_template_dir
        self.output_dir = output_dir or edu_config.report_output_dir

        # 确保目录存在
        Path(self.template_dir).mkdir(parents=True, exist_ok=True)
        Path(self.output_dir).mkdir(parents=True, exist_ok=True)

        # 初始化Jinja2环境
        self.jinja_env = Environment(
            loader=FileSystemLoader(self.template_dir),
            autoescape=select_autoescape(['html', 'xml']),
            trim_blocks=True,
            lstrip_blocks=True
        )

        # 注册自定义过滤器
        self._register_filters()

        # 预定义报告模板
        self.templates = {
            'stem_analysis': ReportTemplate(
                name='STEM能力分析报告',
                template_file='stem_analysis_template.html',
                output_format='pdf',
                charts_enabled=True,
                detailed_stats=True
            ),
            'regional_comparison': ReportTemplate(
                name='区域教育对比报告',
                template_file='regional_comparison_template.html',
                output_format='pdf',
                charts_enabled=True,
                detailed_stats=True
            ),
            'trend_analysis': ReportTemplate(
                name='教育趋势分析报告',
                template_file='trend_analysis_template.html',
                output_format='pdf',
                charts_enabled=True,
                detailed_stats=True
            ),
            'executive_summary': ReportTemplate(
                name='教育数据摘要报告',
                template_file='executive_summary_template.html',
                output_format='pdf',
                charts_enabled=False,
                detailed_stats=False
            )
        }

        # 创建默认模板
        self._create_default_templates()

    def _register_filters(self):
        """注册Jinja2自定义过滤器"""
        def format_percentage(value: float) -> str:
            return f"{value:.1f}%"

        def format_score(value: float) -> str:
            return f"{value:.2f}"

        def subject_display_name(subject: str) -> str:
            subject_names = {
                'math': '数学',
                'science': '科学',
                'technology': '技术',
                'engineering': '工程',
                'language': '语言',
                'arts': '艺术',
                'social_studies': '社会科学',
                'physical_education': '体育'
            }
            return subject_names.get(subject.lower(), subject)

        def grade_display_name(grade: str) -> str:
            grade_names = {
                'elementary': '小学',
                'middle': '初中',
                'high': '高中',
                'university': '大学'
            }
            return grade_names.get(grade.lower(), grade)

        self.jinja_env.filters['percentage'] = format_percentage
        self.jinja_env.filters['score'] = format_score
        self.jinja_env.filters['subject_name'] = subject_display_name
        self.jinja_env.filters['grade_name'] = grade_display_name

    def _create_default_templates(self):
        """创建默认报告模板"""
        templates = {
            'stem_analysis_template.html': self._get_stem_analysis_template(),
            'regional_comparison_template.html': self._get_regional_comparison_template(),
            'trend_analysis_template.html': self._get_trend_analysis_template(),
            'executive_summary_template.html': self._get_executive_summary_template()
        }

        for filename, content in templates.items():
            template_path = Path(self.template_dir) / filename
            if not template_path.exists():
                with open(template_path, 'w', encoding='utf-8') as f:
                    f.write(content)
                logger.info(f"创建默认模板: {filename}")

    def generate_report(self, request: EduReportRequest,
                       data: Dict[str, Any]) -> EduReportMetadata:
        """
        生成教育分析报告

        Args:
            request: 报告请求
            data: 报告数据

        Returns:
            报告元数据
        """
        try:
            # 验证模板
            if request.report_type not in self.templates:
                raise ValueError(f"不支持的报告类型: {request.report_type}")

            template = self.templates[request.report_type]

            # 准备报告数据
            report_data = self._prepare_report_data(data, request)

            # 生成图表
            if request.include_charts and template.charts_enabled:
                chart_paths = self._generate_charts(report_data, request.report_type)
                report_data['chart_paths'] = chart_paths

            # 渲染模板
            html_content = self._render_template(template.template_file, report_data)

            # 生成输出文件
            if request.format.lower() == 'pdf':
                output_path = self._generate_pdf(html_content, request)
            elif request.format.lower() == 'excel':
                output_path = self._generate_excel(report_data, request)
            else:
                raise ValueError(f"不支持的输出格式: {request.format}")

            # 创建报告元数据
            file_size = os.path.getsize(output_path) if os.path.exists(output_path) else 0
            page_count = self._get_page_count(output_path, request.format)

            report_metadata = EduReportMetadata(
                training_id=request.training_id,
                report_type=request.report_type,
                format=request.format.lower(),
                generated_by="system",
                file_path=str(output_path),
                file_size=file_size,
                page_count=page_count,
                data_sources=["federated_learning_results"],
                privacy_level=EduDataPrivacyLevel.HIGH
            )

            logger.info(f"教育报告生成完成: {report_metadata.report_id}")
            return report_metadata

        except Exception as e:
            logger.error(f"生成报告失败: {e}")
            raise

    def _prepare_report_data(self, raw_data: Dict[str, Any],
                           request: EduReportRequest) -> Dict[str, Any]:
        """准备报告数据"""
        prepared_data = {
            'report_title': self._get_report_title(request.report_type),
            'generated_at': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            'training_id': request.training_id,
            'report_period': self._get_report_period(request),
            'summary_stats': self._calculate_summary_stats(raw_data),
            'detailed_data': raw_data if request.include_detailed_stats else {},
            'comparison_regions': request.comparison_regions or [],
            'grade_filter': [g.value for g in request.grade_filter] if request.grade_filter else []
        }

        # 添加特定报告类型的数据
        if request.report_type == 'stem_analysis':
            prepared_data.update(self._prepare_stem_data(raw_data))
        elif request.report_type == 'regional_comparison':
            prepared_data.update(self._prepare_regional_data(raw_data))
        elif request.report_type == 'trend_analysis':
            prepared_data.update(self._prepare_trend_data(raw_data))

        return prepared_data

    def _generate_charts(self, data: Dict[str, Any], report_type: str) -> Dict[str, str]:
        """生成图表"""
        chart_paths = {}

        try:
            # 设置图表样式
            plt.style.use('seaborn-v0_8')

            if report_type == 'stem_analysis':
                # STEM学科得分柱状图
                chart_path = self._create_stem_chart(data)
                chart_paths['stem_scores'] = chart_path

                # 各年级表现对比
                grade_chart_path = self._create_grade_comparison_chart(data)
                chart_paths['grade_comparison'] = grade_chart_path

            elif report_type == 'regional_comparison':
                # 区域对比图表
                regional_chart_path = self._create_regional_chart(data)
                chart_paths['regional_comparison'] = regional_chart_path

            elif report_type == 'trend_analysis':
                # 趋势分析图表
                trend_chart_path = self._create_trend_chart(data)
                chart_paths['trend_analysis'] = trend_chart_path

            logger.info(f"图表生成完成: {len(chart_paths)} 个图表")

        except Exception as e:
            logger.error(f"生成图表失败: {e}")

        return chart_paths

    def _render_template(self, template_file: str, data: Dict[str, Any]) -> str:
        """渲染HTML模板"""
        try:
            template = self.jinja_env.get_template(template_file)
            return template.render(**data)
        except Exception as e:
            logger.error(f"模板渲染失败: {e}")
            # 返回简单的HTML作为后备
            return self._get_simple_html_backup(data)

    def _generate_pdf(self, html_content: str, request: EduReportRequest) -> Path:
        """生成PDF报告"""
        try:
            # 配置PDF选项
            options = {
                'page-size': 'A4',
                'margin-top': '0.75in',
                'margin-right': '0.75in',
                'margin-bottom': '0.75in',
                'margin-left': '0.75in',
                'encoding': "UTF-8",
                'no-outline': None,
                'enable-local-file-access': None
            }

            # 生成文件名
            filename = f"edu_report_{request.training_id}_{request.report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            output_path = Path(self.output_dir) / filename

            # 生成PDF
            pdfkit.from_string(html_content, str(output_path), options=options)

            return output_path

        except Exception as e:
            logger.error(f"PDF生成失败: {e}")
            raise

    def _generate_excel(self, data: Dict[str, Any], request: EduReportRequest) -> Path:
        """生成Excel报告"""
        try:
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "教育数据分析报告"

            # 添加报告标题
            ws['A1'] = self._get_report_title(request.report_type)
            ws['A1'].font = Font(size=16, bold=True)

            # 添加摘要统计
            self._add_summary_to_excel(ws, data.get('summary_stats', {}))

            # 添加详细数据
            if request.include_detailed_stats and 'detailed_data' in data:
                self._add_detailed_data_to_excel(wb, data['detailed_data'])

            # 保存文件
            filename = f"edu_report_{request.training_id}_{request.report_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            output_path = Path(self.output_dir) / filename
            wb.save(output_path)

            return output_path

        except Exception as e:
            logger.error(f"Excel生成失败: {e}")
            raise

    def _create_stem_chart(self, data: Dict[str, Any]) -> str:
        """创建STEM学科得分图表"""
        try:
            stem_scores = data.get('stem_scores', {})
            if not stem_scores:
                return ""

            subjects = list(stem_scores.keys())
            scores = list(stem_scores.values())

            plt.figure(figsize=(10, 6))
            bars = plt.bar(subjects, scores, color=['#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4'])

            plt.title('STEM学科能力得分分布', fontsize=16, pad=20)
            plt.xlabel('学科', fontsize=12)
            plt.ylabel('平均得分', fontsize=12)
            plt.ylim(0, 100)

            # 添加数值标签
            for bar, score in zip(bars, scores):
                plt.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 1,
                        f'{score:.1f}', ha='center', va='bottom')

            plt.xticks(rotation=45)
            plt.tight_layout()

            # 保存图表
            chart_filename = f"stem_chart_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
            chart_path = Path(self.output_dir) / chart_filename
            plt.savefig(chart_path, dpi=300, bbox_inches='tight')
            plt.close()

            return str(chart_path)

        except Exception as e:
            logger.error(f"创建STEM图表失败: {e}")
            return ""

    def _add_summary_to_excel(self, worksheet, summary_stats: Dict[str, Any]):
        """添加摘要统计到Excel"""
        row = 3
        worksheet.cell(row=row, column=1, value="摘要统计").font = Font(bold=True)
        row += 1

        for key, value in summary_stats.items():
            worksheet.cell(row=row, column=1, value=key)
            worksheet.cell(row=row, column=2, value=str(value))
            row += 1

    def _add_detailed_data_to_excel(self, workbook, detailed_data: Dict[str, Any]):
        """添加详细数据到Excel"""
        for sheet_name, data in detailed_data.items():
            if isinstance(data, pd.DataFrame):
                ws = workbook.create_sheet(title=sheet_name[:31])  # Excel限制sheet名称长度
                for r in dataframe_to_rows(data, index=False, header=True):
                    ws.append(r)

    def _get_report_title(self, report_type: str) -> str:
        """获取报告标题"""
        titles = {
            'stem_analysis': 'STEM学科能力分析报告',
            'regional_comparison': '区域教育水平对比报告',
            'trend_analysis': '教育发展态势分析报告',
            'executive_summary': '教育数据执行摘要报告'
        }
        return titles.get(report_type, '教育数据分析报告')

    def _get_report_period(self, request: EduReportRequest) -> str:
        """获取报告周期"""
        if request.time_range:
            start, end = request.time_range
            return f"{start.strftime('%Y-%m-%d')} 至 {end.strftime('%Y-%m-%d')}"
        return f"截至 {datetime.now().strftime('%Y-%m-%d')}"

    def _calculate_summary_stats(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """计算摘要统计"""
        stats = {}

        # 基础统计
        if 'stem_scores' in data:
            stem_scores = data['stem_scores']
            stats['总平均分'] = sum(stem_scores.values()) / len(stem_scores) if stem_scores else 0
            stats['最高分学科'] = max(stem_scores.items(), key=lambda x: x[1])[0] if stem_scores else '无'
            stats['最低分学科'] = min(stem_scores.items(), key=lambda x: x[1])[0] if stem_scores else '无'

        if 'participation_rate' in data:
            stats['参与率'] = f"{data['participation_rate']:.1f}%"

        if 'data_quality_score' in data:
            stats['数据质量评分'] = f"{data['data_quality_score']:.1f}/100"

        return stats

    def _get_page_count(self, file_path: Path, format_type: str) -> Optional[int]:
        """获取页数"""
        if format_type.lower() == 'pdf':
            try:
                # 简单估算PDF页数（假设每页约包含这些内容）
                file_size_kb = file_path.stat().st_size / 1024
                return max(1, int(file_size_kb / 50))  # 估算每页50KB
            except:
                return None
        return None

    def _get_simple_html_backup(self, data: Dict[str, Any]) -> str:
        """简单HTML后备模板"""
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="utf-8">
            <title>{data.get('report_title', '教育报告')}</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 40px; }}
                h1 {{ color: #333; }}
                .stats {{ background: #f5f5f5; padding: 20px; margin: 20px 0; }}
            </style>
        </head>
        <body>
            <h1>{data.get('report_title', '教育数据分析报告')}</h1>
            <p>生成时间: {data.get('generated_at', '未知')}</p>

            <div class="stats">
                <h2>摘要统计</h2>
                <ul>
        """

        for key, value in data.get('summary_stats', {}).items():
            html += f"<li><strong>{key}:</strong> {value}</li>"

        html += """
                </ul>
            </div>
        </body>
        </html>
        """

        return html

    # 模板内容定义
    def _get_stem_analysis_template(self) -> str:
        return """
<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <title>{{ report_title }}</title>
    <style>
        body { font-family: 'Microsoft YaHei', Arial, sans-serif; margin: 0; padding: 40px; }
        .header { text-align: center; margin-bottom: 40px; }
        .header h1 { color: #2c3e50; margin-bottom: 10px; }
        .header p { color: #7f8c8d; }
        .section { margin: 30px 0; }
        .section h2 { color: #34495e; border-bottom: 2px solid #3498db; padding-bottom: 10px; }
        .stats-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 20px; margin: 20px 0; }
        .stat-card { background: #ecf0f1; padding: 20px; border-radius: 8px; text-align: center; }
        .stat-value { font-size: 24px; font-weight: bold; color: #2980b9; }
        .stat-label { color: #7f8c8d; margin-top: 5px; }
        .chart-container { text-align: center; margin: 30px 0; }
        img { max-width: 100%; height: auto; }
        table { width: 100%; border-collapse: collapse; margin: 20px 0; }
        th, td { border: 1px solid #bdc3c7; padding: 12px; text-align: left; }
        th { background-color: #3498db; color: white; }
        tr:nth-child(even) { background-color: #f8f9fa; }
    </style>
</head>
<body>
    <div class="header">
        <h1>{{ report_title }}</h1>
        <p>生成时间: {{ generated_at }} | 训练ID: {{ training_id }}</p>
        <p>报告周期: {{ report_period }}</p>
    </div>

    <div class="section">
        <h2>📊 摘要统计</h2>
        <div class="stats-grid">
            {% for label, value in summary_stats.items() %}
            <div class="stat-card">
                <div class="stat-value">{{ value }}</div>
                <div class="stat-label">{{ label }}</div>
            </div>
            {% endfor %}
        </div>
    </div>

    {% if chart_paths.stem_scores %}
    <div class="section">
        <h2>📈 STEM学科能力分布</h2>
        <div class="chart-container">
            <img src="{{ chart_paths.stem_scores }}" alt="STEM学科得分图表">
        </div>
    </div>
    {% endif %}

    {% if detailed_data.stem_scores %}
    <div class="section">
        <h2>📋 详细学科得分</h2>
        <table>
            <thead>
                <tr>
                    <th>学科</th>
                    <th>平均得分</th>
                    <th>排名</th>
                </tr>
            </thead>
            <tbody>
                {% for subject, score in detailed_data.stem_scores|dictsort(by='value', reverse=true) %}
                <tr>
                    <td>{{ subject|subject_name }}</td>
                    <td>{{ score|score }}</td>
                    <td>{{ loop.index }}</td>
                </tr>
                {% endfor %}
            </tbody>
        </table>
    </div>
    {% endif %}

    <div class="section">
        <h2>ℹ️ 报告说明</h2>
        <p>本报告基于联邦学习框架生成，采用差分隐私技术保护数据安全。</p>
        <p>所有分析结果均为聚合统计，不包含任何个体识别信息。</p>
    </div>
</body>
</html>
        """

    def _get_regional_comparison_template(self) -> str:
        # 类似的模板结构...
        return "<!-- 区域对比报告模板 -->"

    def _get_trend_analysis_template(self) -> str:
        # 类似的模板结构...
        return "<!-- 趋势分析报告模板 -->"

    def _get_executive_summary_template(self) -> str:
        # 类似的模板结构...
        return "<!-- 执行摘要报告模板 -->"

    # 其他图表生成方法的占位符
    def _create_grade_comparison_chart(self, data: Dict[str, Any]) -> str:
        return ""

    def _create_regional_chart(self, data: Dict[str, Any]) -> str:
        return ""

    def _create_trend_chart(self, data: Dict[str, Any]) -> str:
        return ""

    def _prepare_stem_data(self, raw_data: Dict[str, Any]) -> Dict[str, Any]:
        return {"stem_analysis": True}

    def _prepare_regional_data(self, raw_data: Dict[str, Any]) -> Dict[str, Any]:
        return {"regional_analysis": True}

    def _prepare_trend_data(self, raw_data: Dict[str, Any]) -> Dict[str, Any]:
        return {"trend_analysis": True}


# 全局报告生成器实例
report_generator = EduReportGenerator()

if __name__ == "__main__":
    # 测试报告生成
    generator = EduReportGenerator()

    # 测试数据
    test_data = {
        "stem_scores": {
            "math": 85.5,
            "science": 82.3,
            "technology": 78.9,
            "engineering": 80.1
        },
        "participation_rate": 95.2,
        "data_quality_score": 92.8
    }

    # 测试请求
    test_request = EduReportRequest(
        training_id="test_training_001",
        report_type="stem_analysis",
        format="pdf",
        include_charts=True,
        include_detailed_stats=True
    )

    try:
        metadata = generator.generate_report(test_request, test_data)
        print(f"报告生成成功: {metadata.file_path}")
        print(f"文件大小: {metadata.file_size} bytes")
    except Exception as e:
        print(f"报告生成失败: {e}")
